# app/services/export_service.py
from __future__ import annotations
from typing import List, Dict, Iterable, Any, Optional, Tuple, Union
from io import BytesIO
from datetime import date, datetime
import unicodedata
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# TYPE IMPORTS (may be used for attributes; fallback to Any if import fails)
try:
    from ..models import Applicant, ApplicantDoc, ChecklistItem  # type: ignore
except Exception:
    Applicant = Any
    ApplicantDoc = Any
    ChecklistItem = Any

DOC_PREFIX = "doc_"

# ---------- Helpers ----------
def _parse_to_date(v: Optional[object]) -> Optional[date]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None

def _norm_gender(v: Optional[object]) -> str:
    if v in (None, ""):
        return ""
    s = str(v).strip().lower()
    if s in {"1", "m", "male", "nam"}:
        return "Nam"
    if s in {"0", "f", "female", "nu", "nữ", "nư"}:
        return "Nữ"
    if s in {"other", "khac", "khác"}:
        return "Khác"
    return s.capitalize()

def _autosize(ws):
    ws.freeze_panes = "A2"
    for col in ws.columns:
        lengths = [len(str(c.value)) if c.value is not None else 0 for c in col]
        w = max(10, *(lengths or [10])) + 2
        ws.column_dimensions[col[0].column_letter].width = min(w, 40)

def _get(obj: Union[dict, Any], key: str, default=None):
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)

def _display_name_from_obj(obj: Union[dict, Any]) -> str:
    hd = (_get(obj, "ho_dem") or "").strip()
    t = (_get(obj, "ten") or "").strip()
    if hd or t:
        return f"{hd} {t}".strip()
    return (_get(obj, "ho_ten") or "").strip()

def _split_name_cells(obj: Union[dict, Any]) -> Tuple[str, str]:
    ln = (_get(obj, "ho_dem") or "").strip()
    fn = (_get(obj, "ten") or "").strip()
    if ln or fn:
        return ln, fn
    full = (_get(obj, "ho_ten") or "").strip()
    if not full:
        return "", ""
    parts = full.split()
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]

def _set_date_format_by_header(ws, headers: List[str], header_names: Iterable[str]):
    name_to_idx = {h: i + 1 for i, h in enumerate(headers)}  # 1-based
    for hn in header_names:
        col = name_to_idx.get(hn)
        if not col:
            continue
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                if isinstance(c.value, (date, datetime)):
                    c.number_format = "dd/mm/yyyy"
                    c.alignment = Alignment(horizontal="center")

# ---------- Normalization helpers ----------
def _normalize_text(s: Optional[str]) -> str:
    """Lowercase, remove diacritics, remove non-alphanumeric (keep letters+digits)."""
    if not s:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]", "", s)
    return s

# ---------- Document allocation rules ----------
# Use the exact column names you provided; normalization + fuzzy will handle variants
SPECIAL_DOCS = [
    "Bằng tốt nghiệp Đại học",
    "Bảng điểm toàn khoá học Đại học",
    "Bằng tốt nghiệp Cao đẳng",
    "Bảng điểm toàn khóa học Cao đẳng",
    "Bằng tốt nghiệp Trung Cấp",
    "Bảng điểm toàn khóa Trung Cấp",
]
EXEMPT_DOC = "Đơn miễn giảm"

_SPECIAL_DOCS_N = {_normalize_text(x) for x in SPECIAL_DOCS}
_EXEMPT_DOC_N = _normalize_text(EXEMPT_DOC)

def _is_special_item(it: Any) -> bool:
    code = getattr(it, "code", None)
    disp = getattr(it, "display_name", None)
    return (_normalize_text(code) in _SPECIAL_DOCS_N) or (_normalize_text(disp) in _SPECIAL_DOCS_N)

def _is_exempt_item(it: Any) -> bool:
    code = getattr(it, "code", None)
    disp = getattr(it, "display_name", None)
    return (_normalize_text(code) == _EXEMPT_DOC_N) or (_normalize_text(disp) == _EXEMPT_DOC_N)

def _get_item_qty_from_dm(dm: Dict[str, int], it: Any) -> int:
    """
    Robust lookup of qty for item `it` in dm (mapping keys -> qty).
    Try (in order):
      1) exact dm[it.code]
      2) exact dm[it.display_name]
      3) normalized exact of code/display_name
      4) fuzzy substring match between normalized keys
    """
    if not dm:
        return 0

    code = getattr(it, "code", None)
    disp = getattr(it, "display_name", None)

    # direct exact matches
    try:
        if code is not None and code in dm:
            return int(dm.get(code, 0) or 0)
        if disp is not None and disp in dm:
            return int(dm.get(disp, 0) or 0)
    except Exception:
        pass

    code_n = _normalize_text(code)
    disp_n = _normalize_text(disp)

    # normalized exact
    try:
        if code_n and code_n in dm:
            return int(dm.get(code_n, 0) or 0)
        if disp_n and disp_n in dm:
            return int(dm.get(disp_n, 0) or 0)
    except Exception:
        pass

    # fuzzy: try match against any dm key normalized
    try:
        # build normalized map of dm keys to qty (cache)
        for k, v in dm.items():
            kn = _normalize_text(k)
            if not kn:
                continue
            # if normalized dm key is substring of item normalized OR vice versa
            if (code_n and (kn in code_n or code_n in kn)) or (disp_n and (kn in disp_n or disp_n in kn)):
                return int(v or 0)
    except Exception:
        pass

    # fallback 0
    return 0

def split_doc_rows(dm: Dict[str, int], items_all: List[ChecklistItem]) -> Tuple[List[int], List[int]]: # type: ignore
    """
    For one applicant's dm (code->qty) and ordered items_all returns:
      - main_row: numbers for 'Hồ sơ Nhập học'
      - reduced_row: numbers for 'Hồ sơ miễn giảm'
    Rules:
      - SPECIAL_DOCS: main = 1 if qty>0 else 0; reduced = qty - main
      - EXEMPT_DOC: main = 0; reduced = qty
      - other: main = qty; reduced = 0
    Keep zeros as integers.
    """
    main_row: List[int] = []
    reduced_row: List[int] = []

    for it in items_all or []:
        qty = _get_item_qty_from_dm(dm or {}, it)
        if _is_special_item(it):
            main_qty = 1 if qty > 0 else 0
            reduced_qty = qty - main_qty if qty > 0 else 0
        elif _is_exempt_item(it):
            main_qty = 0
            reduced_qty = qty
        else:
            main_qty = qty
            reduced_qty = 0

        main_row.append(main_qty)
        reduced_row.append(reduced_qty)

    return main_row, reduced_row

# ---------- Export builder ----------
def build_excel_bytes_by_items(
    apps: List[Applicant], # type: ignore
    docs: List[ApplicantDoc], # type: ignore
    items: List[ChecklistItem], # type: ignore
    *,
    split_name: bool = False,
) -> bytes:
    """
    Build Excel with two sheets:
      - 'Hồ sơ Nhập học' (main_row)
      - 'Hồ sơ miễn giảm' (reduced_row)
    """
    # Build mapping applicant_ma_so_hv -> {key_variant -> qty}
    docs_by_mssv: Dict[str, Dict[str, int]] = {}
    for d in docs or []:
        m = docs_by_mssv.setdefault(d.applicant_ma_so_hv, {})
        key = d.code
        # store original key
        m[key] = int(d.so_luong or 0)
        # store normalized key too for faster direct lookup
        key_n = _normalize_text(key)
        if key_n:
            # only set normalized key if not already present (preserve original if duplicate)
            if key_n not in m:
                m[key_n] = int(d.so_luong or 0)

    base_headers = [
        "STT", "Mã hồ sơ", "Ngày nhận", "Email học viên"
    ]
    if split_name:
        base_headers += ["Họ và tên", "Họ đệm", "Tên"]
    else:
        base_headers += ["Họ và tên"]

    base_headers += [
        "MSHV", "Ngày sinh", "Số ĐT", "Ngành nhập học", "Đợt", "Khóa",
        "Đã TN trước đó", "Ghi chú", "Người nhận", "Dân tộc"
    ]

    item_headers = [getattr(it, "display_name", None) or it.code for it in (items or [])]
    headers = base_headers + item_headers

    wb = Workbook()

    # Sheet 1: Hồ sơ Nhập học
    ws = wb.active
    ws.title = "Hồ sơ Nhập học"
    ws.append(headers)

    for idx, a in enumerate(apps or [], start=1):
        common_prefix = [
            idx,
            a.ma_ho_so or "",
            _parse_to_date(getattr(a, "ngay_nhan_hs", None)),
            a.email_hoc_vien or "",
        ]

        if split_name:
            full = _display_name_from_obj(a)
            ln, fn = _split_name_cells(a)
            name_cells = [full, ln, fn]
        else:
            name_cells = [_display_name_from_obj(a)]

        common_suffix = [
            a.ma_so_hv or "",
            _parse_to_date(getattr(a, "ngay_sinh", None)),
            a.so_dt or "",
            getattr(a, "nganh_nhap_hoc", None) or getattr(a, "nganh", None) or "",
            a.dot or "",
            getattr(a, "khoa", "") or "",
            a.da_tn_truoc_do or "",
            a.ghi_chu or "",
            a.nguoi_nhan_ky_ten or "",
            getattr(a, "dan_toc", None) or "",
        ]

        dm = docs_by_mssv.get(a.ma_so_hv, {})

        main_row, _reduced_row = split_doc_rows(dm, items or [])

        # Keep zeros as integers
        row = common_prefix + name_cells + common_suffix + main_row
        ws.append(row)

    # Sheet 2: Hồ sơ miễn giảm
    ws2 = wb.create_sheet("Hồ sơ miễn giảm")
    ws2.append(headers)

    for idx, a in enumerate(apps or [], start=1):
        common_prefix = [
            idx,
            a.ma_ho_so or "",
            _parse_to_date(getattr(a, "ngay_nhan_hs", None)),
            a.email_hoc_vien or "",
        ]

        if split_name:
            full = _display_name_from_obj(a)
            ln, fn = _split_name_cells(a)
            name_cells = [full, ln, fn]
        else:
            name_cells = [_display_name_from_obj(a)]

        common_suffix = [
            a.ma_so_hv or "",
            _parse_to_date(getattr(a, "ngay_sinh", None)),
            a.so_dt or "",
            getattr(a, "nganh_nhap_hoc", None) or getattr(a, "nganh", None) or "",
            a.dot or "",
            getattr(a, "khoa", "") or "",
            a.da_tn_truoc_do or "",
            a.ghi_chu or "",
            a.nguoi_nhan_ky_ten or "",
            getattr(a, "dan_toc", None) or "",
        ]

        dm = docs_by_mssv.get(a.ma_so_hv, {})
        _main_row, reduced_row = split_doc_rows(dm, items or [])

        row = common_prefix + name_cells + common_suffix + reduced_row
        ws2.append(row)

    # Freeze & autosize both sheets
    for ws_sheet in (ws, ws2):
        ws_sheet.freeze_panes = "A2"
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws_sheet[letter]:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws_sheet.column_dimensions[letter].width = min(max(10, max_len + 2), 40)

    # Set date formats for date columns
    _set_date_format_by_header(ws, headers, header_names=["Ngày nhận", "Ngày sinh"])
    _set_date_format_by_header(ws2, headers, header_names=["Ngày nhận", "Ngày sinh"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()